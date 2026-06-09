"""
build.py — NPFR Strategic Plan build script
Reads the latest data/statuses_*.json as the single source of truth.
Run before git push:
    python build.py
Outputs:
    data/index.json
    exports/NPFR_Strategic_Plan_YYYYMMDD.xlsx
"""
import json, glob
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE    = Path(__file__).parent
DATA    = BASE / "data"
EXPORTS = BASE / "exports"
EXPORTS.mkdir(exist_ok=True)

# ── Brand palette ──────────────────────────────────────────────────────────
HEX = {
    "bg":        "1C0505", "edge":      "7B1416", "edge_dark": "4A0A0C",
    "gold":      "C8A040", "gold_lt":   "E8C870", "cream":     "F5E6C8",
    "subtitle":  "A07850", "light_row": "FFF5E6", "alt_row":   "FAEBD7",
    "complete":  "0d3320", "comp_txt":  "4ade80",
    "inprog":    "2a2a0a", "inprog_txt":"d4d46a",
    "notstart":  "1a1a2e", "ns_txt":    "8888aa",
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(h="F5E6C8", bold=False, size=10, italic=False):
    return Font(color=h, bold=bold, size=size, italic=italic, name="Calibri")
def ctr(wrap=False): return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def lft(wrap=True):  return Alignment(horizontal="left",   vertical="top",    wrap_text=wrap, indent=1)
def side(s="thin", c="C8A040"): return Side(style=s, color=c)
def bdr(): return Border(left=side(), right=side(), top=side(), bottom=side())

STATUS_FMT = {
    "Complete":    (HEX["complete"],  HEX["comp_txt"]),
    "In Progress": (HEX["inprog"],    HEX["inprog_txt"]),
    "Not Started": (HEX["notstart"],  HEX["ns_txt"]),
}
TASK_FMT = {
    "complete":    (HEX["complete"],  HEX["comp_txt"]),
    "in-progress": (HEX["inprog"],    HEX["inprog_txt"]),
    "not-started": (HEX["notstart"],  HEX["ns_txt"]),
}
TASK_LBL = {"complete":"Complete","in-progress":"In Progress","not-started":"Not Started"}

# ── Load latest plan JSON ──────────────────────────────────────────────────
def load_plan():
    files = sorted(DATA.glob("statuses_*.json"), reverse=True)
    if not files:
        raise FileNotFoundError("No statuses_*.json files found in data/")
    with open(files[0]) as f:
        return json.load(f), files[0]

# ── Helpers ────────────────────────────────────────────────────────────────
def all_goals(plan):
    return [g for p in plan["pillars"] for g in p["goals"]]

def gates(plan):
    g = {}
    for goal in all_goals(plan):
        if goal.get("predecessor"):
            g.setdefault(goal["predecessor"], []).append(goal["id"])
    return g

def goal_status(g):
    s = [t["status"] for t in g["tasks"]]
    if all(x == "complete"    for x in s): return "Complete"
    if all(x == "not-started" for x in s): return "Not Started"
    return "In Progress"

def task_counts(plan):
    total = done = 0
    for p in plan["pillars"]:
        for g in p["goals"]:
            for t in g["tasks"]:
                total += 1
                if t["status"] == "complete": done += 1
    return total, done

def pillar_counts(p):
    total = done = 0
    for g in p["goals"]:
        for t in g["tasks"]:
            total += 1
            if t["status"] == "complete": done += 1
    return total, done

def sc(ws, row, col, val, fnt=None, fil=None, aln=None):
    c = ws.cell(row=row, column=col, value=val)
    if fnt: c.font = fnt
    if fil: c.fill = fil
    if aln: c.alignment = aln
    return c

def mc(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

# ── Cover sheet ────────────────────────────────────────────────────────────
def build_cover(wb, plan):
    ws = wb.active; ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 20
    for r in range(1, 52):
        for col in "ABC": ws[f"{col}{r}"].fill = fill(HEX["edge_dark"])
    for col in "ABC": ws[f"{col}4"].fill = fill(HEX["gold"])
    ws.row_dimensions[6].height = 42
    mc(ws,6,2,6,3)
    sc(ws,6,2,"NORTH PORT FIRE RESCUE",
       fnt=Font(name="Calibri",bold=True,size=28,color=HEX["cream"]),
       aln=Alignment(horizontal="left",vertical="center"))
    ws.row_dimensions[8].height = 28
    mc(ws,8,2,8,3)
    sc(ws,8,2,f"Strategic Plan  ·  FY2026 – FY2029",
       fnt=Font(name="Calibri",size=18,color=HEX["gold"]),
       aln=Alignment(horizontal="left",vertical="center"))
    for col in "ABC": ws[f"{col}10"].fill = fill(HEX["gold"])
    for label, row, val in [
        ("MISSION",12,plan["mission"]),
        ("VISION", 15,plan.get("vision","")),
        ("VALUES", 18," · ".join(plan.get("values",[]))),
    ]:
        ws.row_dimensions[row].height = 13
        ws.row_dimensions[row+1].height = 28
        sc(ws,row,2,label, fnt=Font(name="Calibri",bold=True,size=8,color=HEX["gold"]),
           aln=Alignment(horizontal="left",vertical="bottom"))
        mc(ws,row+1,2,row+1,3)
        sc(ws,row+1,2,val, fnt=Font(name="Calibri",size=10,color=HEX["cream"]),
           aln=Alignment(horizontal="left",vertical="center",wrap_text=True))
    hr = 22
    ws.row_dimensions[hr].height = 18
    mc(ws,hr,2,hr,3)
    sc(ws,hr,2,"STRATEGIC PILLARS",
       fnt=Font(name="Calibri",bold=True,size=10,color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=ctr())
    for i, p in enumerate(plan["pillars"]):
        r = hr+1+i; ws.row_dimensions[r].height = 22
        total, done = pillar_counts(p)
        pct = f"{round(done/total*100)}%" if total else "0%"
        sc(ws,r,2,p["id"],
           fnt=Font(name="Calibri",bold=True,size=10,color=HEX["gold"]),
           fil=fill(HEX["edge"]), aln=ctr())
        sc(ws,r,3,f"{p['name']}  —  Steward: {p['steward']}  ({done}/{total} tasks, {pct})",
           fnt=Font(name="Calibri",size=10,color=HEX["cream"]),
           fil=fill(HEX["edge_dark"]),
           aln=Alignment(horizontal="left",vertical="center",indent=1))
    fr = hr+len(plan["pillars"])+3
    mc(ws,fr,2,fr,3)
    sc(ws,fr,2,f"Generated {datetime.now().strftime('%B %d, %Y')}  ·  {plan.get('label','')}",
       fnt=Font(name="Calibri",size=8,color=HEX["subtitle"],italic=True),
       aln=Alignment(horizontal="left",vertical="center"))

# ── Summary sheet ──────────────────────────────────────────────────────────
def build_summary(wb, plan):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    widths = [2,10,40,18,8,8,8,10]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width = w
    for col in range(1,9): ws.cell(row=1,column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 24
    mc(ws,2,2,2,8)
    sc(ws,2,2,"STRATEGIC PLAN SUMMARY",
       fnt=Font(name="Calibri",bold=True,size=14,color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=ctr())
    ws.row_dimensions[3].height = 18
    for col,lbl in enumerate(["Pillar","Name","Steward","Goals","Tasks","Done","Complete %"],start=2):
        sc(ws,3,col,lbl,
           fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),
           fil=fill(HEX["edge"]), aln=ctr())
    gt = gd = 0
    for i,p in enumerate(plan["pillars"]):
        r = 4+i; ws.row_dimensions[r].height = 18
        total, done = pillar_counts(p)
        gt += total; gd += done
        bg = fill(HEX["alt_row"] if i%2 else HEX["light_row"])
        pct_val = done/total if total else 0
        sc(ws,r,2,p["id"],fnt=Font(name="Calibri",bold=True,size=9,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
        sc(ws,r,3,p["name"],fnt=font("1a1a1a",size=9),fil=bg,aln=Alignment(horizontal="left",vertical="center",indent=1))
        sc(ws,r,4,p["steward"],fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        sc(ws,r,5,len(p["goals"]),fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        sc(ws,r,6,total,fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        sc(ws,r,7,done,fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        c = sc(ws,r,8,pct_val,fnt=Font(name="Calibri",bold=True,size=9,color="1a1a1a"),fil=bg,aln=ctr())
        c.number_format = "0%"
    r = 4+len(plan["pillars"]); ws.row_dimensions[r].height = 18
    mc(ws,r,2,r,5)
    sc(ws,r,2,"TOTAL",fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    sc(ws,r,6,gt,fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    sc(ws,r,7,gd,fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    c = sc(ws,r,8,gd/gt if gt else 0,fnt=Font(name="Calibri",bold=True,size=9,color=HEX["gold"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    c.number_format = "0%"

# ── Year view sheet ────────────────────────────────────────────────────────
def build_year_sheet(wb, plan):
    ws = wb.create_sheet("By Year")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40
    for col in range(1,8): ws.cell(row=1,column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 24
    mc(ws,2,2,2,7)
    sc(ws,2,2,"GOALS BY FISCAL YEAR",
       fnt=Font(name="Calibri",bold=True,size=14,color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=ctr())
    years = sorted(set(g["year"] for p in plan["pillars"] for g in p["goals"]))
    row = 3
    for yr in years:
        ws.row_dimensions[row].height = 20
        mc(ws,row,2,row,7)
        sc(ws,row,2,yr,
           fnt=Font(name="Calibri",bold=True,size=11,color=HEX["cream"]),
           fil=fill(HEX["edge"]), aln=ctr())
        row += 1
        ws.row_dimensions[row].height = 16
        for col,lbl in enumerate(["Goal ID","Pillar","Goal Statement","Owner","Status","Success Metric"],start=2):
            sc(ws,row,col,lbl,
               fnt=Font(name="Calibri",bold=True,size=8,color=HEX["cream"]),
               fil=fill(HEX["edge_dark"]), aln=ctr())
        row += 1
        yr_goals = [(p,g) for p in plan["pillars"] for g in p["goals"] if g["year"]==yr]
        for i,(p,g) in enumerate(yr_goals):
            ws.row_dimensions[row].height = 28
            gs = goal_status(g)
            gs_bg, gs_fg = STATUS_FMT[gs]
            bg = fill(HEX["alt_row"] if i%2 else HEX["light_row"])
            sc(ws,row,2,g["id"],fnt=Font(name="Calibri",bold=True,size=9,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
            sc(ws,row,3,p["name"][:18],fnt=font("1a1a1a",size=8),fil=bg,aln=Alignment(horizontal="left",vertical="center",indent=1))
            sc(ws,row,4,g["statement"],fnt=font("1a1a1a",size=9),fil=bg,aln=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1))
            sc(ws,row,5,g.get("owner",""),fnt=font("3a3a5c",size=8,italic=True),fil=bg,aln=ctr())
            sc(ws,row,6,gs,fnt=Font(name="Calibri",bold=True,size=8,color=gs_fg),fil=fill(gs_bg),aln=ctr())
            sc(ws,row,7,g["metric"],fnt=font("1a1a1a",size=8),fil=bg,aln=Alignment(horizontal="left",vertical="top",wrap_text=True,indent=1))
            row += 1
        row += 1  # gap between years

# ── Pillar sheets ──────────────────────────────────────────────────────────
def build_pillar_sheet(wb, p, gate_map):
    ws = wb.create_sheet(f"{p['id']} {p['name'][:20]}")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJK", [2,10,8,14,12,45,18,6,50,14,42]):
        ws.column_dimensions[col].width = w
    for col in range(1,12): ws.cell(row=1,column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    mc(ws,2,2,2,11)
    sc(ws,2,2,f"{p['id']}  —  {p['name']}",
       fnt=Font(name="Calibri",bold=True,size=14,color=HEX["edge_dark"]),
       fil=fill(HEX["gold"]), aln=Alignment(horizontal="left",vertical="center",indent=1))
    ws.row_dimensions[3].height = 16
    mc(ws,3,2,3,11)
    sc(ws,3,2,f"Steward: {p['steward']}",
       fnt=Font(name="Calibri",size=9,italic=True,color=HEX["cream"]),
       fil=fill(HEX["edge"]), aln=Alignment(horizontal="left",vertical="center",indent=1))
    ws.row_dimensions[4].height = 18
    for col,lbl in enumerate(["Goal ID","Year","Status","Predecessor","Goal Statement","Owner","T#","Task","Task Status","Success Metric"],start=2):
        sc(ws,4,col,lbl, fnt=Font(name="Calibri",bold=True,size=8,color=HEX["cream"]),
           fil=fill(HEX["edge_dark"]), aln=ctr())
    # Group goals by year within pillar
    years = sorted(set(g["year"] for g in p["goals"]))
    row = 5
    for yr in years:
        yr_goals = [g for g in p["goals"] if g["year"]==yr]
        ws.row_dimensions[row].height = 16
        mc(ws,row,2,row,11)
        sc(ws,row,2,yr, fnt=Font(name="Calibri",bold=True,size=10,color=HEX["edge_dark"]),
           fil=fill(HEX["gold"]), aln=Alignment(horizontal="left",vertical="center",indent=1))
        row += 1
        for gi,g in enumerate(yr_goals):
            gs = goal_status(g)
            gs_bg, gs_fg = STATUS_FMT[gs]
            n = len(g["tasks"])
            gate_marker = " ★" if g["id"] in gate_map else ""
            for tr in range(n):
                rr = row+tr; ws.row_dimensions[rr].height = 18
                bg = HEX["alt_row"] if gi%2 else HEX["light_row"]
                if tr == 0:
                    mc(ws,rr,2,rr+n-1,2)
                    sc(ws,rr,2,f"{g['id']}{gate_marker}", fnt=Font(name="Calibri",bold=True,size=9,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
                    mc(ws,rr,3,rr+n-1,3)
                    sc(ws,rr,3,g["year"],fnt=font("1a1a1a",size=9),fil=fill(bg),aln=ctr())
                    mc(ws,rr,4,rr+n-1,4)
                    sc(ws,rr,4,gs,fnt=Font(name="Calibri",bold=True,size=9,color=gs_fg),fil=fill(gs_bg),aln=ctr())
                    mc(ws,rr,5,rr+n-1,5)
                    sc(ws,rr,5,g.get("predecessor") or "—",fnt=font("1a1a1a",size=9),fil=fill(bg),aln=ctr())
                    mc(ws,rr,6,rr+n-1,6)
                    sc(ws,rr,6,g["statement"],fnt=font("1a1a1a",size=9),fil=fill(bg),aln=Alignment(horizontal="left",vertical="top",wrap_text=True,indent=1))
                    mc(ws,rr,7,rr+n-1,7)
                    sc(ws,rr,7,g.get("owner",""),fnt=font("3a3a5c",size=9,italic=True),fil=fill(bg),aln=Alignment(horizontal="left",vertical="center",indent=1))
                    mc(ws,rr,11,rr+n-1,11)
                    sc(ws,rr,11,g["metric"],fnt=font("1a1a1a",size=8,italic=True),fil=fill(bg),aln=Alignment(horizontal="left",vertical="top",wrap_text=True,indent=1))
                t = g["tasks"][tr]
                ts_bg, ts_fg = TASK_FMT[t["status"]]
                sc(ws,rr,8,f"T{tr+1}",fnt=Font(name="Calibri",bold=True,size=8,color=HEX["cream"]),fil=fill(HEX["edge"]),aln=ctr())
                sc(ws,rr,9,t["text"],fnt=font("1a1a1a",size=9),fil=fill(bg),aln=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1))
                sc(ws,rr,10,TASK_LBL[t["status"]],fnt=Font(name="Calibri",bold=True,size=8,color=ts_fg),fil=fill(ts_bg),aln=ctr())
            for col in range(2,12):
                ws.cell(row=row+n,column=col).border = Border(top=Side(style="thin",color=HEX["gold"]))
            row += n
        row += 1  # year gap

# ── Dependencies sheet ─────────────────────────────────────────────────────
def build_dep_sheet(wb, plan, gate_map):
    ws = wb.create_sheet("Dependencies")
    ws.sheet_view.showGridLines = False
    for col,w in zip("ABCDE",[2,14,55,14,55]):
        ws.column_dimensions[col].width = w
    for col in range(1,6): ws.cell(row=1,column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 24
    mc(ws,2,2,2,5)
    sc(ws,2,2,"DEPENDENCY CHAINS",fnt=Font(name="Calibri",bold=True,size=14,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
    ws.row_dimensions[3].height = 16
    mc(ws,3,2,3,5)
    sc(ws,3,2,"★ = Gate goal — must complete before downstream goals can begin",
       fnt=Font(name="Calibri",size=9,italic=True,color=HEX["cream"]),
       fil=fill(HEX["edge"]),aln=Alignment(horizontal="left",vertical="center",indent=1))
    ws.row_dimensions[4].height = 16
    for col,lbl in enumerate(["Gate Goal","Gate Statement","Unlocks","Unlocked Statement"],start=2):
        sc(ws,4,col,lbl,fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    all_g = {g["id"]:g for p in plan["pillars"] for g in p["goals"]}
    row = 5
    for gate_id, deps in gate_map.items():
        gate_g = all_g.get(gate_id)
        if not gate_g: continue
        for di, dep_id in enumerate(deps):
            dep_g = all_g.get(dep_id)
            if not dep_g: continue
            ws.row_dimensions[row].height = 28
            bg = fill(HEX["alt_row"] if row%2 else HEX["light_row"])
            sc(ws,row,2,f"{gate_id} ★",fnt=Font(name="Calibri",bold=True,size=9,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
            sc(ws,row,3,gate_g["statement"],fnt=font("1a1a1a",size=9),fil=bg,aln=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1))
            sc(ws,row,4,dep_id,fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
            sc(ws,row,5,dep_g["statement"],fnt=font("1a1a1a",size=9),fil=bg,aln=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1))
            row += 1

# ── Changelog sheet ────────────────────────────────────────────────────────
def build_changelog_sheet(wb, plan):
    if not plan.get("changelog"): return
    ws = wb.create_sheet("Changelog")
    ws.sheet_view.showGridLines = False
    for col,w in zip("ABCDE",[2,22,14,14,60]):
        ws.column_dimensions[col].width = w
    for col in range(1,6): ws.cell(row=1,column=col).fill = fill(HEX["edge_dark"])
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 24
    mc(ws,2,2,2,5)
    sc(ws,2,2,"CHANGELOG",fnt=Font(name="Calibri",bold=True,size=14,color=HEX["edge_dark"]),fil=fill(HEX["gold"]),aln=ctr())
    ws.row_dimensions[3].height = 16
    for col,lbl in enumerate(["Timestamp","Type","Goal","Detail"],start=2):
        sc(ws,3,col,lbl,fnt=Font(name="Calibri",bold=True,size=9,color=HEX["cream"]),fil=fill(HEX["edge_dark"]),aln=ctr())
    for i,entry in enumerate(reversed(plan["changelog"])):
        r = 4+i; ws.row_dimensions[r].height = 20
        bg = fill(HEX["alt_row"] if i%2 else HEX["light_row"])
        sc(ws,r,2,entry.get("timestamp",""),fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        sc(ws,r,3,entry.get("type",""),fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        sc(ws,r,4,entry.get("goalId",""),fnt=font("1a1a1a",size=9),fil=bg,aln=ctr())
        note = entry.get("note") or f"{entry.get('taskText','')[:60]}  {entry.get('from','')} → {entry.get('to','')}"
        sc(ws,r,5,note,fnt=font("1a1a1a",size=9),fil=bg,aln=Alignment(horizontal="left",vertical="center",wrap_text=True,indent=1))

# ── Rebuild index.json ─────────────────────────────────────────────────────
def rebuild_index():
    files = sorted([f.name for f in DATA.glob("statuses_*.json")], reverse=True)
    idx = {"files":files,"latest":files[0] if files else None,
           "generated":datetime.now().isoformat(timespec="seconds")}
    with open(DATA/"index.json","w") as f: json.dump(idx,f,indent=2)
    print(f"  index.json → {len(files)} file(s), latest: {idx['latest']}")

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    print("NPFR Strategic Plan — build.py")
    print("="*42)
    print("\n[1] Rebuilding data/index.json")
    rebuild_index()
    print("\n[2] Loading plan data")
    plan, src = load_plan()
    print(f"  Source: {src.name}")
    print(f"  Pillars: {len(plan['pillars'])}  Goals: {sum(len(p['goals']) for p in plan['pillars'])}  Tasks: {sum(len(g['tasks']) for p in plan['pillars'] for g in p['goals'])}")
    gate_map = gates(plan)
    print(f"  Gate goals: {len(gate_map)}")
    print("\n[3] Building Excel workbook")
    wb = Workbook()
    build_cover(wb, plan)
    build_summary(wb, plan)
    build_year_sheet(wb, plan)
    for p in plan["pillars"]:
        build_pillar_sheet(wb, p, gate_map)
        print(f"  {p['id']}  {p['name']}")
    build_dep_sheet(wb, plan, gate_map)
    build_changelog_sheet(wb, plan)
    date_str = datetime.now().strftime("%Y%m%d")
    out = EXPORTS / f"NPFR_Strategic_Plan_{date_str}.xlsx"
    wb.save(out)
    print(f"\n  Saved: exports/NPFR_Strategic_Plan_{date_str}.xlsx")
    print("\n[4] Completion summary")
    total, done = task_counts(plan)
    for p in plan["pillars"]:
        pt, pd = pillar_counts(p)
        bar = "█"*int(pd/pt*20)+"░"*(20-int(pd/pt*20)) if pt else "░"*20
        print(f"  {p['id']}  {bar}  {pd}/{pt}")
    print(f"\n  Overall: {done}/{total} tasks complete ({round(done/total*100) if total else 0}%)")
    print(f"\nDone. Ready to push.\n")

if __name__ == "__main__":
    main()